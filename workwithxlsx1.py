import openpyxl

wb = openpyxl.load_workbook('test2.xlsx')
sheet = wb['Лист1']
print(sheet['c2'].value)
c = sheet['b4']
print(c.column)
print(c.row)
print(c.coordinate)
d = sheet.cell(row=3, column=4)
print(d.value, d.coordinate)
print('*' * 50)
wb = openpyxl.load_workbook('tab.xlsx')
sh = wb.active
print(sh.values)
nmin = sh.cell(row=7, column=2).value
for rownum in range(8, 28):
    nmin = min(nmin, sh.cell(row=rownum, column=2).value)
print(nmin)


# print(wb.sheetnames)
# shet  = wb.sheetnames[0]
# print(help(openpyxl.load_workbook))